import schedule  
import time 
import threading
from datetime import datetime
from mainDB import MainDB
import csv
from tkinter import *

import random


from publish import PublishFeed
from login import ParlerLogin
from unfollowing import ParlerUnfollow
from listOfFollowers import ListOfFollowers


import os, sys
from datetime import datetime
import openpyxl

class SettingsPage():

	def createSettingsPage(self):

		followModule = self.tk.Label(self.tab, text="Daily automation",
		 foreground="white",
		 background="black", 
		 width=47,
		 height=1
			)
		followModule.config(font=("Courier", 15))
		followModule.grid(row=0, columnspan=2)


		followModuleCronLabel= self.tk.Label(self.tab,text="Run automation every (n) hour (example : 15:00,16:00...) for Follow module\n (separated by commas)")
		followModuleCronLabel.grid(row=1, columnspan=2, padx=10, pady=10)
		self.followModuleCron = self.tk.Text(self.tab, font = ('courier', 12, 'bold'), width=45, height=4)
		self.followModuleCron.grid(row=2, columnspan=2)

		followCronCredRow = MainDB.followModuleCron()
		followModuleCronCredString = ""
		for item in followCronCredRow:
			followModuleCronCredString += item + ","
		self.followModuleCron.insert(self.tk.END, followModuleCronCredString[:-1])



		unfollowModuleCronLabel= self.tk.Label(self.tab,text="Run automation every (n) hour (example : 15:00,16:00...) for Unfollow module\n (separated by commas)")
		unfollowModuleCronLabel.grid(row=3, columnspan=2, padx=10, pady=10)
		self.unfollowModuleCron = self.tk.Text(self.tab, font = ('courier', 12, 'bold'), width=45, height=4)
		self.unfollowModuleCron.grid(row=4, columnspan=2)

		unfollowCronCredRow = MainDB.unFollowModuleCron()
		unfollowModuleCronCredString = ""
		for item in unfollowCronCredRow:
			unfollowModuleCronCredString += item + ","
		self.unfollowModuleCron.insert(self.tk.END, unfollowModuleCronCredString[:-1])


		publishModuleCronLabel= self.tk.Label(self.tab,text="Run automation every (n) hour (example : 15:00,16:00...) for Post Scheduler module\n (separated by commas)")
		publishModuleCronLabel.grid(row=5, columnspan=2, padx=10, pady=10)
		self.publishModuleCron = self.tk.Text(self.tab, font = ('courier', 12, 'bold'), width=45, height=4)
		self.publishModuleCron.grid(row=6, columnspan=2)

		publishSchCronCredRow = MainDB.postSchCron()
		publishSchCronCredString = ""
		for item in publishSchCronCredRow:
			publishSchCronCredString += item + ","
		self.publishModuleCron.insert(self.tk.END, publishSchCronCredString[:-1])

		followLimitLabel = self.tk.Label(self.tab, text="Daily follow limit")
		followLimitLabel.grid(column=0, row=7, padx=10, pady=10)

		storedFollowLimit = StringVar(self.tab, value=MainDB.followLimitCron()[0])
		self.followLimit = self.tk.Entry(self.tab, textvariable=storedFollowLimit, font = ('courier', 12, 'bold'))
		self.followLimit.grid(column=0, row=8, pady=10)


		unfollowLimitLabel = self.tk.Label(self.tab, text="Daily unfollow limit")
		unfollowLimitLabel.grid(column=1, row=7, padx=10, pady=10)

		storedUnfollowLimit = StringVar(self.tab, value=MainDB.unFollowLimitCron()[0])
		self.unFollowLimit = self.tk.Entry(self.tab, textvariable=storedUnfollowLimit, font = ('courier', 12, 'bold'))
		self.unFollowLimit.grid(column=1, row=8, pady=10)



		self.saveAllSettings = self.tk.Button(self.tab, text ="Save All", command = self.saveSettingDataWithReload)
		self.saveAllSettings.grid(row=10, column=0,  padx=10, pady=10)

		self.followModuleOn = self.tk.Button(self.tab, text ="Start Follow Module", command = self.startOnlyFollowModule)
		self.followModuleOn.grid(row=10, column=1,  padx=10, pady=10)

		self.unfollowModuleOn = self.tk.Button(self.tab, text ="Start Unfollow Module", command = self.startOnlyUnfollowModule)
		self.unfollowModuleOn.grid(row=11, column=0,  padx=10, pady=10)

		self.publishModuleOn = self.tk.Button(self.tab, text ="Start Publish Module", command = self.startOnlyPublishModule)
		self.publishModuleOn.grid(row=11, column=1,  padx=10, pady=10)

		self.startAutomation = self.tk.Button(self.tab, text ="Start automation", command = self.startCronTread)
		self.startAutomation.grid(row=12, column=0,  padx=10, pady=10)

		self.stopAutomation = self.tk.Button(self.tab, text ="Stop automation", command = self.stopExecutingCronJob)
		self.stopAutomation.grid(row=12, column=1,  padx=10, pady=10)
		

		# self.test = self.tk.Button(self.tab, text ="Test", command = self.test2323)
		# self.test.grid(row=30, columnspan=2,  padx=10, pady=10)


		taskManager = self.tk.Label(self.tab, text="Task Manager",
		 foreground="white",
		 background="black", 
		 width=47,
		 height=1
			)
		taskManager.config(font=("Courier", 15))
		taskManager.grid(row=13, columnspan=2)

		self.CronTask = self.tk.Label(self.tab,text='Automation : off')
		self.CronTask.grid(row=14, columnspan=2, padx=10, pady=10)
		self.CronTask.config(bg= "#c4143d", fg= "#fff")

		self.followCronTask = self.tk.Label(self.tab,text="Follow module : off")
		self.followCronTask.grid(row=15, columnspan=2, padx=10, pady=10)
		self.followCronTask.config(bg= "#c4143d", fg= "#fff")

		self.unfollowCronTask = self.tk.Label(self.tab,text="Unfollow module : off")
		self.unfollowCronTask.grid(row=16, columnspan=2, padx=10, pady=10)
		self.unfollowCronTask.config(bg= "#c4143d", fg= "#fff")

		self.publishCronTask = self.tk.Label(self.tab,text="Publish module : off")
		self.publishCronTask.grid(row=17, columnspan=2, padx=10, pady=10)
		self.publishCronTask.config(bg= "#c4143d", fg= "#fff")


	def saveCronData(self):
		openDBCsv = open('database/followModuleCronCSV.csv', 'w')
		writerCSV = csv.writer(openDBCsv)
		writerCSV.writerow(self.followModuleCron.get('1.0', 'end-1c').split(","))
		openDBCsv.close()

		openDBCsv = open('database/unFollowModuleCronCSV.csv', 'w')
		writerCSV = csv.writer(openDBCsv)
		writerCSV.writerow(self.unfollowModuleCron.get('1.0', 'end-1c').split(","))
		openDBCsv.close()

		openDBCsv = open('database/postSchCronCSV.csv', 'w')
		writerCSV = csv.writer(openDBCsv)
		writerCSV.writerow(self.publishModuleCron.get('1.0', 'end-1c').split(","))
		openDBCsv.close()

		followLimitArr = []
		followLimitArr.append(self.followLimit.get())
		openDBCsv = open('database/followLimitCronCSV.csv', 'w')
		writerCSV = csv.writer(openDBCsv)
		writerCSV.writerow(followLimitArr)
		openDBCsv.close()

		unfollowLimitArr = []
		unfollowLimitArr.append(self.unFollowLimit.get())
		openDBCsv = open('database/unfollowLimitCronCSV.csv', 'w')
		writerCSV = csv.writer(openDBCsv)
		writerCSV.writerow(unfollowLimitArr)
		openDBCsv.close()

		# t7 = threading.Thread(target=self.startCron);
		# t7.start()



	def startCronTread(self):
		print(self.loginCred)
		print(self.badgesDB)
		self.CronTask.config(text = 'Automation : on')
		self.CronTask.config(bg= "#19a62e", fg= "#212121")

		t5 = threading.Thread(target=self.startCron);
		t5.start()

	# def test123(self):
	# 	followLimitArr = []
	# 	followLimitArr.append(self.followLimit.get())
	# 	openDBCsv = open('database/followLimitCronCSV.csv', 'w')
	# 	writerCSV = csv.writer(openDBCsv)
	# 	writerCSV.writerow(followLimitArr)
	# 	openDBCsv.close()

	# 	t7 = threading.Thread(target=self.test12);
	# 	t7.start()

	# def test12(self):
	# 	self.job = schedule.every().day.at('18:16').do(self.stopCronTread)
	# 	self.job = schedule.every().day.at('18:18').do(self.stopCronTread)
	# 	self.job = schedule.every().day.at('18:20').do(self.stopCronTread)

	# 	while True:
	# 		self.sch = schedule.run_pending()
	# 		time.sleep(1)

	def stopExecutingCronJob(self):
		schedule.clear()
		self.CronTask.config(text = 'Automation : off')
		self.CronTask.config(bg= "#c4143d", fg= "#fff")

		self.followCronTask.config(text = "Follow module : off")
		self.followCronTask.config(bg= "#c4143d", fg= "#fff")

		self.unfollowCronTask.config(text = "Unfollow module : off")
		self.unfollowCronTask.config(bg= "#c4143d", fg= "#fff")

		self.publishCronTask.config(text = "Publish module : off")
		self.publishCronTask.config(bg= "#c4143d", fg= "#fff")

		python = sys.executable
		os.execl(python, python, * sys.argv)

	def startExecutingCronJobFollow_Module(self):

		self.followCronTask.config(text = "Follow module : on")
		self.followCronTask.config(bg= "#19a62e", fg= "#212121")

		now = datetime.now()
		current_time = now.strftime("%d-%b")
		if(os.path.exists(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")):
			if (os.access(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx",os.R_OK)):
				wrkbk = openpyxl.load_workbook(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")
				sh = wrkbk.active
				numberOfExcelRows = sh.max_row
				wrkbk.save(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")
				wrkbk.close()
			else:
				print("File permision denied: Follow module")
		else:
			book = openpyxl.Workbook()
			book.save(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")
			book.close()
			newEmptyRow = 1
			numberOfExcelRows = 0
		

		if(os.access(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx",os.R_OK)):
			if(int(numberOfExcelRows) < int(self.followLimit.get())):
				ParlerLogin("Proccessing", self.followModuleHash, self.badgesDB, self.loginCred, self.passCred)
		else:
			print("File permision denied: Follow module")

		self.followCronTask.config(text = "Follow module : off")
		self.followCronTask.config(bg= "#c4143d", fg= "#fff")


		time.sleep(2)

		
	def startExecutingCronJobUnfollow_Module(self):
		# ################################################################
		# UNFOLLOW MODULE CRON

		self.unfollowCronTask.config(text = "Unfollow module : on")
		self.unfollowCronTask.config(bg= "#19a62e", fg= "#212121")

		now = datetime.now()
		current_time = now.strftime("%d-%b")

		if(os.path.exists(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")):
			if (os.access(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx",os.R_OK)):
				wrkbk = openpyxl.load_workbook(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")
				sh = wrkbk.active
				numberOfExcelRows = sh.max_row
				wrkbk.save(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")
				wrkbk.close()
			else:
				print("File permision denied: Unfollow module")
		else:
			book = openpyxl.Workbook()
			book.save(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")
			book.close()
			newEmptyRow = 1
			numberOfExcelRows = 0
		
		if (os.access(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx",os.R_OK)):
			if(int(numberOfExcelRows) < int(self.unFollowLimit.get())):
				ParlerUnfollow(self.loginCred, self.passCred, self.unfollowException, self.unFollowLimit.get())
		else:
			print("File permision denied: Unfollow module")


		self.unfollowCronTask.config(text = "Unfollow module : off")
		self.unfollowCronTask.config(bg= "#c4143d", fg= "#fff")

		time.sleep(2)

	def startExecutingCronJobPublish_Module(self):
		# ################################################################
		# PUBLISH MODULE CRON

		self.publishCronTask.config(text = "Publish module : on")
		self.publishCronTask.config(bg= "#19a62e", fg= "#212121")

		if (os.access(os.path.abspath("") + "\\" + "publishedFeed.xlsx",os.R_OK)):
			PublishFeed(self.loginCred, self.passCred, self.postSch, self.postSchNum, self.postSchHash)
		else:
			print("File permision denied: Publish module")

		self.publishCronTask.config(text = "Publish module : off")
		self.publishCronTask.config(bg= "#c4143d", fg= "#fff")

		time.sleep(2)


	# def job_that_executes_once(self):
	# 	now = datetime.now()
	# 	current_time = now.strftime("%H:%M:%S")
	# 	print("test: "+current_time)

	# def job_that_executes_once_publishModule(self):
	# 	PublishFeed(self.loginCred, self.passCred, self.postSch, self.postSchNum, self.postSchHash)


	def createListOfFollowers(self):
		ListOfFollowers(self.loginCred, self.passCred, self.unfollowException, self.unFollowLimit.get())

	def startCron(self):

		self.saveCronData()

		# print(str(self.followModuleCron.get('1.0', 'end-1c')).split(","))
		# followArr = self.createCronJobForModule(str(self.followModuleCron.get('1.0', 'end-1c')).split(","), [[6,11],[40,42]])
		followArr = self.createCronJobForModule(str(self.followModuleCron.get('1.0', 'end-1c')).split(","), [[53,53],[44,44]])

		# self.job_that_executes_once_publishModule()

		print("Follow Module Time")
		print(followArr)

		unfollowArr = self.createCronJobForModule(str(self.unfollowModuleCron.get('1.0', 'end-1c')).split(","), [[15,21]])

		print("Unfollow Module Time")
		print(unfollowArr)

		# publishArr = self.createCronJobForModule(str(self.publishModuleCron.get('1.0', 'end-1c')).split(","), [[29,32]])
		publishArr = self.createCronJobForModule(str(self.publishModuleCron.get('1.0', 'end-1c')).split(","), [[58,58]])

		print("Publish Module Time")
		print(publishArr)

		self.job = schedule.every().day.at('01:02').do(self.createListOfFollowers)

		for x in followArr:
			self.job = schedule.every().day.at(x).do(self.startExecutingCronJobFollow_Module)

		for x2 in unfollowArr:
			self.job = schedule.every().day.at(x2).do(self.startExecutingCronJobUnfollow_Module)

		for x3 in publishArr:
			self.job = schedule.every().day.at(x3).do(self.startExecutingCronJobPublish_Module)
			

		# self.job = schedule.every().day.at('13:08').do(self.job_that_executes_once)

		# self.job = schedule.every().day.at('13:10').do(self.job_that_executes_once)


		while True:
			self.sch = schedule.run_pending()
			time.sleep(1)


	def createCronJobForModule(self, module1, timeTest):
		numberBetween = []
		inc = -1
		for x1 in timeTest:
			inc = inc + 1
			for x in range(int(module1[0]),int(module1[1])):
				# print(inc)
				timeForUpdate = str(x)
				if timeForUpdate in ['0','1','2','3','4','5','6','7','8','9']:
					timeForUpdate = "0"+timeForUpdate

				randomTime = str(random.randint(timeTest[inc][0],timeTest[inc][1]))
				if randomTime in ['0','1','2','3','4','5','6','7','8','9']:
					randomTime = "0"+randomTime

				numberBetween.append(str(timeForUpdate)+":"+randomTime)

		return numberBetween


	def __init__(self, window, tab, tk, loginCred, passCred, postSch, postSchNum, postSchHash, followModuleHash, badgesDB, unfollowException):
		self.window = window
		self.tab = tab
		self.tk = tk
		self.job = ""
		self.startAutomation = ""
		self.stopAutomation = ""
		self.saveAllSettings = ""
		self.followModuleOn = ""
		self.unfollowModuleOn = ""
		self.publishModuleOn = ""

		self.followModuleCron = ""
		self.unfollowModuleCron = ""
		self.publishModuleCron = ""
		self.followLimit = ""
		self.unFollowLimit = ""

		# Task manager box
		self.CronTask = ""
		self.followCronTask = ""
		self.unfollowCronTask = ""
		self.publishCronTask = ""


		self.loginCred = loginCred
		self.passCred = passCred

		# Follow module
		self.followModuleHash = followModuleHash
		self.badgesDB = badgesDB

		# Unfollow module
		self.unfollowException = unfollowException

		# Publish feed
		self.postSch = postSch
		self.postSchNum = postSchNum
		self.postSchHash = postSchHash



		self.createSettingsPage()


	# #####################################################################
	# ADDITIONAL FEATURES
	# #####################################################################

	# ############### START FOLLOW MODULE
	def saveSettingDataWithReload(self):
		schedule.clear()
		self.saveCronData()
		python = sys.executable
		os.execl(python, python, * sys.argv)


	# ############### START FOLLOW MODULE
	def startOnlyFollowModule(self):
		t35 = threading.Thread(target=self.startCronFollow);
		t35.start()

	def startCronFollow(self):
		self.saveCronData()

		followArr = self.createCronJobForModule(str(self.followModuleCron.get('1.0', 'end-1c')).split(","), [[6,11],[40,42]])

		print("Follow Module Time")
		print(followArr)

		for x in followArr:
			self.job = schedule.every().day.at(x).do(self.startExecutingCronJobFollow_Module)


		while True:
			self.sch = schedule.run_pending()
			time.sleep(1)


	# ############### START UNFOLLOW MODULE
	def startOnlyUnfollowModule(self):
		t36 = threading.Thread(target=self.startCronUnfollow);
		t36.start()

	def startCronUnfollow(self):
		self.saveCronData()

		# unfollowArr = self.createCronJobForModule(str(self.unfollowModuleCron.get('1.0', 'end-1c')).split(","), [[17,25],[30,37]])
		unfollowArr = self.createCronJobForModule(str(self.unfollowModuleCron.get('1.0', 'end-1c')).split(","), [[27,27]])

		print("Unfollow Module Time")
		print(unfollowArr)

		for x2 in unfollowArr:
			self.job = schedule.every().day.at(x2).do(self.startExecutingCronJobUnfollow_Module)

		while True:
			self.sch = schedule.run_pending()
			time.sleep(1)

	# ############### START PUBLISH MODULE
	def startOnlyPublishModule(self):
		t37 = threading.Thread(target=self.startCronPublish);
		t37.start()

	def startCronPublish(self):
		self.saveCronData()

		# unfollowArr = self.createCronJobForModule(str(self.unfollowModuleCron.get('1.0', 'end-1c')).split(","), [[17,25],[30,37]])
		# publishArr = self.createCronJobForModule(str(self.publishModuleCron.get('1.0', 'end-1c')).split(","), [[29,32],[55,59]])
		publishArr = self.createCronJobForModule(str(self.publishModuleCron.get('1.0', 'end-1c')).split(","), [[16,16],[15,20]])
		print("Publish Module Time")
		print(publishArr)

		for x3 in publishArr:
			self.job = schedule.every().day.at(x3).do(self.startExecutingCronJobPublish_Module)

		while True:
			self.sch = schedule.run_pending()
			time.sleep(1)

