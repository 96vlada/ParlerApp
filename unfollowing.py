import time
import xlsxwriter

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import TimeoutException
import random

import threading
import os, sys

from datetime import datetime, timedelta
import openpyxl


class ParlerUnfollow:
	def checkForUnfollow(self):

		# five days fix

		fiveDaysOldFollowing = []
		for x in range(6):
			# if(x == 0):
			# 	continue
			dateForIter = datetime.today() - timedelta(days=x)
			iterDate = dateForIter.strftime("%d-%b")
			
			if(os.path.exists(os.path.abspath("archive") + "\\" + "following"+iterDate+".xlsx")):
				# print(os.path.abspath("archive") + "\\" + "unfollowing"+iterDate+".xlsx")
				wrkbk = openpyxl.load_workbook(os.path.abspath("archive") + "\\" + "following"+iterDate+".xlsx")
				sh = wrkbk.active
				newEmptyRow = sh.max_row
				# print(newEmptyRow)
				for x1 in range(newEmptyRow):
					if(sh.cell(row = x1+1, column = 1).value == None):
						continue
					fiveDaysOldFollowing.append(sh.cell(row = x1+1, column = 1).value)


		for itemFollowing in fiveDaysOldFollowing:
			self.listOfFollowers.append(itemFollowing)



		# ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
		# CONFIG_PATH = os.path.join(ROOT_DIR, 'chromedriver.exe')
		# service = Service(CONFIG_PATH)
		# # service = Service('C:/Users/chromedriver.exe')
		# driver = webdriver.Chrome(service=service)
		# driver.set_window_size(1980, 1020)
		# driver.get("https://parler.com/login.php")
		
		options = Options()
		options.add_argument('--headless')
		options.add_argument('--disable-gpu')
		options.add_argument('--ignore-certificate-errors')
		options.add_argument('--ignore-ssl-errors')

		chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")

		driver = webdriver.Chrome(chromedriver_path, chrome_options=options)
		driver.set_window_size(1980, 1020)
		driver.get("https://parler.com/login.php")
		driver.implicitly_wait(356)
		

		def send_keys_delay(controller,keys,min_delay=0.05,max_delay=0.05):
			for key in keys:
				controller.send_keys(key)
				time.sleep(random.uniform(min_delay,max_delay))

		send_keys_delay(driver.find_element(By.XPATH, '//input[@type="text"]'), self.parlerUsername)
		send_keys_delay(driver.find_element(By.XPATH, '//input[@type="password"]'), self.parlerPassword)
		driver.find_element(By.XPATH, '//button[@class="button"]').click()
		time.sleep(4)

		driver.find_elements_by_xpath('//ul[@class="sidebar__menu"]/li/a')[1].click()
		time.sleep(2)

		driver.find_elements_by_xpath('//p[@class="user-card__follow"]/a')[0].click()
		time.sleep(2)

		if(self.unfollowException[0] != ""):
			for unEx in self.unfollowException:
				self.listOfFollowers.append(unEx)


		# print(list(driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/h3')))
		
		# incFollowers = 0


		################################################################################################

		# allFollowers = []
		########################### 40 MINUTES

		if(os.path.exists(os.path.abspath("") + "\\" + "listOfFollowers.xlsx")):
			if (os.access(os.path.abspath("") + "\\" + "listOfFollowers.xlsx",os.R_OK)):
				wrkbk = openpyxl.load_workbook(os.path.abspath("") + "\\" + "listOfFollowers.xlsx")
				sh = wrkbk.active
				newEmptyRow = sh.max_row
				for x2 in range(newEmptyRow):
					if(sh.cell(row = x2+1, column = 1).value == None):
						continue
					time.sleep(1)
					self.listOfFollowers.append(sh.cell(row = x2+1, column = 1).value)


		# print(allFollowers)


		# for x in range(120):
		# 	driver.execute_script("window.scrollTo(0,1000000)");
		# 	time.sleep(3)
		# time.sleep(3)
		# for e in driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/h3'):
		# 	self.listOfFollowers.append(e.text)
		# 	# incFollowers = incFollowers + 1
		# time.sleep(3)

		# print(self.listOfFollowers);


		################################################################################################


		# driver.execute_script("window.scrollTo(0,0)")



		driver.find_elements_by_xpath('//p[@class="user-card__follow"]/a')[1].click()
		time.sleep(2)

		inc = 0
		incWindowScroll = 0




		for x in range(90):
			time.sleep(4)
			driver.execute_script("window.scrollTo(0,document.body.scrollHeight)")


		# workbook = xlsxwriter.Workbook('unfollowing.xlsx')
		# worksheet = workbook.add_worksheet()
		# row = 0
		# column = 0

		# for e in driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/h3'):
		# driver.execute_script("window.scrollTo(0,0)")

		time.sleep(356)
		testinc = 0;
		

		for e in driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/p'):
			if (testinc > int(self.unfollowLimit)):
				break

			testinc = testinc + 1
			testinc2 = "-" + str(testinc)
			testinc2 = int(testinc2)
			# if(inc > int(self.unfollowLimit)):
			# 	break
			# time.sleep(3)

			time.sleep(3)

			# print(driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/p')[testinc2].text)

			if driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/p')[testinc2].text not in self.listOfFollowers:

				# worksheet.write(row, column, e.text)
				# row += 1
					# print(driver.find_elements_by_xpath('//div[contains(text(), "Unfollow") and @class="button"]')[inc].location.get('y'))
					# driver.execute_script("window.scrollTo(0," + str(driver.find_elements_by_xpath('//div[contains(text(), "Unfollow") and @class="button"]')[inc].location.get('y') - 100) + ")")

				# time.sleep(2)
				#####################################################
				driver.execute_script("window.scrollTo(0," + str(driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/h3')[testinc2].location.get('y') - 100) + ")")
				# driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div[2]')[testinc2].click()
				
				time.sleep(3)
				element2 = driver.find_elements_by_xpath('//div[contains(text(), "Unfollow") and @class="button"]')[testinc2]
				driver.execute_script("arguments[0].click();", element2)

				self.listOFUnfollowing.append(driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/p')[testinc2].text)


				
				# driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div[2]')[testinc2].click()
				#####################################################

				time.sleep(2)
				
				# self.listOfFollowing.append(driver.find_elements_by_xpath('//ul[@class="basic-list"]/li/div/a[2]/p')[testinc2].text)

				inc = inc - 1

				
				# WebDriverWait(driver, 4).until(find).click()
				time.sleep(3)
				# driver.execute_script("window.scrollTo(0," + str(incWindowScroll) + ")")

			inc = inc + 1
			
		# workbook.close()

		self.createExcelForFollowing()
		
		time.sleep(2)


		# print("List of Following" + str(self.listOfFollowing))
		# print("List of Followers" + str(self.listOfFollowers))


		

	def createExcelForFollowing(self):
		now = datetime.now()
		current_time = now.strftime("%d-%b")
		if(os.path.exists(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")):
			if (os.access(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx",os.R_OK)):
				wrkbk = openpyxl.load_workbook(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")
				sh = wrkbk.active
				newEmptyRow = sh.max_row + 1
				wrkbk.save(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")
				wrkbk.close()
			else:
				print("File permision denied: Unfollow module")
		else:
			book = openpyxl.Workbook()
			book.save(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")
			book.close()
			newEmptyRow = 1

		book = openpyxl.load_workbook(filename=os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")
		sheet = book.active

		for item in self.listOFUnfollowing :
			newEmptyRow = newEmptyRow + 1;
			sheet.cell(row=newEmptyRow, column=1).value = item
		book.save(os.path.abspath("archive") + "\\" + "unfollowing"+current_time+".xlsx")
		book.close()


	def __init__(self, parlerUsername, parlerPassowrd, unfollowException, unfollowLimit):
		self.parlerUsername = parlerUsername
		self.parlerPassword = parlerPassowrd
		self.listOfFollowing = []
		self.listOfFollowers = []
		self.unfollowException = unfollowException
		self.unfollowLimit = unfollowLimit
		self.listOFUnfollowing = []

		t2 = threading.Thread(target=self.checkForUnfollow);
		t2.start()

		t2.join()
		print("Unfollowing module completed")