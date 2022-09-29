import time
import xlsxwriter
import feedparser
import csv

from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import NoSuchElementException
import random

import threading
import os, sys
import re
import openpyxl


class PublishFeed:

	def doPublishFeed(self):

		# ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
		# CONFIG_PATH = os.path.join(ROOT_DIR, 'chromedriver.exe')
		# service = Service(CONFIG_PATH)
		# # service = Service('C:/Users/chromedriver.exe')
		# driver = webdriver.Chrome(service=service)
		# driver.set_window_size(1980, 1020)
		# driver.get("https://parler.com/login.php")

		options = Options()
		options.add_argument('--headless')
		options.add_argument('--disable-gpu')
		options.add_argument('--ignore-certificate-errors')
		options.add_argument('--ignore-ssl-errors')
		chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")
		driver = webdriver.Chrome(chromedriver_path, chrome_options=options)
		driver.set_window_size(1980, 1020)
		driver.get("https://parler.com/login.php")

		def send_keys_delay(controller,keys,min_delay=0.05,max_delay=0.05):
			for key in keys:
				controller.send_keys(key)
				time.sleep(random.uniform(min_delay,max_delay))

		send_keys_delay(driver.find_element(By.XPATH, '//input[@type="text"]'), self.parlerUsername)
		send_keys_delay(driver.find_element(By.XPATH, '//input[@type="password"]'), self.parlerPassword)
		driver.find_element(By.XPATH, '//button[@class="button"]').click()

		# feedToExec = ['https://dev-rss-mix.pantheonsite.io/?feed=rss2','https://www.lesalonbeige.fr/feed/']
		feedToExec = self.feeds

		incForPublish = 0
		for feedUrl in feedToExec:
			incForPublish = incForPublish + 1
			feedForPublish = feedparser.parse(feedUrl)


			# print(len(feedForPublish.entries))
			driver.execute_script("window.scrollTo(0," + str(0) + ")")
			for i in range(int(self.feedNumb) - 1,-1,-1):


				time.sleep(3)
				# if(os.path.exists('publishedFeed.csv')):
				# 	with open('publishedFeed.csv', newline='') as f:
				# 		reader = csv.reader(f)
				# 		for badge in f:
				# 			if(badge != "\r\n"):
				# 				self.badgeArr = badge.rstrip().split(",")
				# 				self.badgeArr = list(filter(None, self.badgeArr))

			
				publishedFeedReview = []
				if(os.path.exists(os.path.abspath("") + "\\" + "publishedFeed.xlsx")):
					# print(os.path.abspath("archive") + "\\" + "unfollowing"+iterDate+".xlsx")
					wrkbk = openpyxl.load_workbook(os.path.abspath("") + "\\" + "publishedFeed.xlsx")
					sh = wrkbk.active
					newEmptyRow = sh.max_row
					for x1 in range(newEmptyRow):
						if(sh.cell(row = x1+1, column = 1).value == None):
							continue
						publishedFeedReview.append(sh.cell(row = x1+1, column = 1).value)


				if str(feedForPublish.entries[i].published[5:]) in publishedFeedReview:
					continue
				
				time.sleep(2)
				# driver.find_elements_by_xpath('//ul[@class="sidebar__menu"]/li/a')[1].click()
				# time.sleep(3)
				# driver.find_element_by_xpath('//div[@class="input-trigger"]').click()

				# driver.find_element_by_xpath('//textarea[@class="modal__parley-input"]').click()
				# time.sleep(2)

				# driver.find_element(By.XPATH, '//div[@class="emoji-wysiwyg-editor"]').clear()
				# time.sleep(2)
				entry = feedForPublish.entries[i]

				if(len(entry.title) > 80):
					title_entry = entry.title[0:80] + "..."
				else:
					title_entry = entry.title

				send_keys_delay(driver.find_element(By.XPATH, '//textarea[@class="modal__parley-input"]'), title_entry)
				time.sleep(1)
				driver.find_element(By.XPATH, '//textarea[@class="modal__parley-input"]').send_keys(Keys.ENTER)
				time.sleep(1)
				driver.find_element(By.XPATH, '//textarea[@class="modal__parley-input"]').send_keys(Keys.ENTER)
				time.sleep(1)

				#########################################################################
				# tags
				hashtagsFromFeedTags = ""
				for x in feedForPublish.entries[i].tags:
					tagValidation = x.term
					if (' ' in x.term):
						tagValidation = x.term.replace(" ","-")
					if ('+' in x.term):
						tagValidation = x.term.replace("+","")

					hashtagsFromFeedTags += " #" + tagValidation
				# tags
				#########################################################################


				# #############################################################################
				if(str(entry.description[ 0 : 2 ]) == ""):
					clean_data_without_img = "#"+self.hashtags.replace(",", " #") + " " + entry.link

					clean_data_without_img += hashtagsFromFeedTags

				if(str(entry.description[ 0 : 2 ]) == "<p"):
					clean_data_without_img =  entry.description[3:]
					clean_data_without_img = clean_data_without_img[0:120] + " " + "#"+self.hashtags.replace(",", " #") + " " + entry.link

					clean_data_without_img += hashtagsFromFeedTags

				if(str(entry.description[ 0 : 4 ]) == "<img"):
					clean_data_without_img = re.sub("(<img.*?>)", "", entry.description, 0, re.IGNORECASE | re.DOTALL | re.MULTILINE)[0:150] + "..." + " " + "#"+self.hashtags.replace(",", " #") + " " + entry.link
					
					clean_data_without_img += hashtagsFromFeedTags

					if(str(clean_data_without_img[ 0 : 2 ]) == "<p"):
						clean_data_without_img = clean_data_without_img[3:]

				if(str(entry.description[ 0 : 7 ]) == "<p><img"):
					clean = re.compile('<.*?>')
					clean_data_without_img = re.sub(clean, '', entry.description)[0:120] + "..." + " " + "#"+self.hashtags.replace(",", " #") + " " + entry.link

					clean_data_without_img += hashtagsFromFeedTags
					
				# #############################################################################

				driver.find_element(By.XPATH, '//textarea[@class="modal__parley-input"]').send_keys(clean_data_without_img)
				time.sleep(1)

				driver.find_element(By.XPATH, '//div[@class="modal__footer"]/button[@type="button"]').click()

				time.sleep(3)

				self.publishedFeed.append(entry.published[5:])

				driver.execute_script("window.scrollTo(0," + str(0) + ")")

			time.sleep(3)


		# with open("Quiz.csv", "a", newline="") as file:
		# 	writer = csv.writer(file)
		# 	writer.writerow([Name, "Question", "Solution", "Mark"])

		# 	writer.writerow(self.publishedFeed)
		# 	writer.writerow([2, str(x) + " - " + str(y) , A2, score2])

		# openDBCsvFeed = open('publishedFeed.csv', 'a')
		# writerCSV = csv.writer(openDBCsvFeed)
		# writerCSV.writerow(self.publishedFeed)
		# openDBCsvFeed.close()
		self.createPublishedFeedExcel()


	def createPublishedFeedExcel(self):
		if(os.path.exists(os.path.abspath("") + "\\" + "publishedFeed.xlsx")):
			if (os.access(os.path.abspath("") + "\\" + "publishedFeed.xlsx",os.R_OK)):
				wrkbk = openpyxl.load_workbook(os.path.abspath("") + "\\" + "publishedFeed.xlsx")
				sh = wrkbk.active
				newEmptyRow = sh.max_row
				wrkbk.save(os.path.abspath("") + "\\" + "publishedFeed.xlsx")
				wrkbk.close()
			else:
				print("File permision denied: Publish module")
		else:
			book = openpyxl.Workbook()
			book.save(os.path.abspath("") + "\\" + "publishedFeed.xlsx")
			book.close()
			newEmptyRow = 1

		listOfFollowing = ['test121','test23']
		book = openpyxl.load_workbook(filename=os.path.abspath("") + "\\" + "publishedFeed.xlsx")
		sheet = book.active
		incForRow = 0;
		for item in self.publishedFeed :
			incForRow = incForRow + 1
			sheet.cell(row=newEmptyRow + incForRow, column=1).value = item
		book.save(os.path.abspath("") + "\\" + "publishedFeed.xlsx")
		book.close()



	def __init__(self, parlerUsername, parlerPassowrd, feeds, feedNumb, hashtags):
		self.parlerUsername = parlerUsername
		self.parlerPassword = parlerPassowrd
		self.listOfFollowing = []
		self.listOfFollowers = []
		self.feeds = feeds.split(",")
		self.feedNumb = feedNumb
		self.hashtags = hashtags
		self.publishedFeed = []
		self.badgeArr = []

		t3 = threading.Thread(target=self.doPublishFeed);
		t3.start()

		t3.join()
		print("Publish module completed")