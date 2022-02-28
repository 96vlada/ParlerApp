from tkinter import ttk
import tkinter
import time

import xlsxwriter

# ########################################################################
from selenium import webdriver
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.chrome.service import Service
from selenium.common.exceptions import WebDriverException
from selenium.common.exceptions import NoSuchElementException    
import random

from datetime import datetime
import openpyxl
# from openpyxl.workbook import Workbook

# ########################################################################

import threading
import os, sys



class ParlerLogin:

    def parlerSelenium(self):

        # ROOT_DIR = os.path.dirname(os.path.abspath(__file__))
        # CONFIG_PATH = os.path.join(ROOT_DIR, 'chromedriver.exe')
        # service = Service(CONFIG_PATH)
        # # service = Service('C:/Users/chromedriver.exe')
        # driver = webdriver.Chrome(service=service)
        # driver.set_window_size(1980, 1020)
        # driver.get("https://parler.com/login.php")
        
        options = Options()
        options.add_argument('--headless')
        options.add_argument('--disable-gpu')
        options.add_argument('--ignore-certificate-errors')
        options.add_argument('--ignore-ssl-errors')
        chromedriver_path = os.path.join(sys._MEIPASS, "chromedriver.exe")

        driver = webdriver.Chrome(chromedriver_path, chrome_options=options)
        driver.set_window_size(1980, 1020)
        driver.get("https://parler.com/login.php")

        def send_keys_delay(controller,keys,min_delay=0.05,max_delay=0.05):
            for key in keys:
                controller.send_keys(key)
                time.sleep(random.uniform(min_delay,max_delay))

        send_keys_delay(driver.find_element(By.XPATH, '//input[@type="text"]'), self.parlerUsername)
        send_keys_delay(driver.find_element(By.XPATH, '//input[@type="password"]'), self.parlerPassword)
        driver.find_element(By.XPATH, '//button[@class="button"]').click()
        time.sleep(1)

        hashtagList = self.hashtag.split(",")

        for hashtagName in hashtagList:
            

            if(hashtagName != "none"):

                driver.get('https://parler.com/hashtags/'+hashtagName)

                time.sleep(4)


                try:
                    testname = ""
                    incWindowScroll = 0

                    for inc in range(18):
                        try:
                            if(inc != 0):
                                # print(driver.find_elements_by_xpath('//div[@class="post__author"]/h1[@class="post__header__name"]/a')[inc].location.get('y') - 200)

                                driver.execute_script("window.scrollTo(0," + str(driver.find_elements_by_xpath('//div[@class="post__author"]/h1[@class="post__header__name"]/a')[inc].location.get('y') - 100) + ")")
                                
                                time.sleep(2)
                            # driver.execute_script("window.scrollTo(0," + str(incWindowScroll) + ")")
                            
                            driver.find_elements_by_xpath('//div[@class="post__author"]/h1[@class="post__header__name"]/a')[inc].click()
                            time.sleep(2)

                            try:
                                time.sleep(2)
                                if "all" not in self.badges:
                                    for badge in driver.find_elements_by_class_name('badges__badge'):
                                        for badgeInner in self.badges:
                                            # print(str(badgeInner))
                                            if(badge.get_attribute("alt") == str(badgeInner)):
                                                if(str(badgeInner) != "all"):
                                                    followBtn = driver.find_element_by_xpath('//div[@class="user-card__buttons"]/button')
                                                    if(followBtn.get_attribute("title")[ 0 : 6 ] == "Follow"):
                                                        followBtn.click()

                                                        self.userNames.append(driver.find_element_by_xpath('//div[@class="user-card__user-details"]/h1[@class="name"]').text)

                                                        # self.createExcelForFollowing()

                                                        time.sleep(1)

                                                        
                                else:
                                    followBtn = driver.find_element_by_xpath('//div[@class="user-card__buttons"]/button')
                                    if(followBtn.get_attribute("title")[ 0 : 6 ] == "Follow"):
                                        followBtn.click()

                                        self.userNames.append(driver.find_element_by_xpath('//div[@class="user-card__user-details"]/h1[@class="name"]').text)

                                        # self.createExcelForFollowing()

                                        time.sleep(1)


                                            # print(badge.get_attribute("alt"))

                            except NoSuchElementException:
                                print("No element found")

                            time.sleep(2)


                            # incWindowScroll = incWindowScroll + 500


                            time.sleep(2)    
                            driver.back()
                            time.sleep(2)    
                        except WebDriverException:
                            print("Is not clickable")
                        else:
                            print("Is not clickable second try")


                    time.sleep(2)    

                except IndexError:
                    gotdata = 'null'

     

                time.sleep(2)

            # if hashtag is none ####################################
            else:
                # time.sleep(2)
                # driver.find_elements_by_xpath('//ul[@class="sidebar__menu"]/li/a')[3].click()

                time.sleep(4)

                noneInc = 0
                noneIncFix = 0

                for user in driver.find_elements_by_xpath('//ul[@class="user-list"]/li[@class="user-list__item"]/button[@class="button"]'):
                    noneIncFix = noneIncFix + 1
                    # if(noneInc == 20):
                    if(noneIncFix == 24):
                        break

                    if(noneInc == 6):
                        time.sleep(3)
                        driver.refresh()
                        noneInc = 0
                        time.sleep(3)


                    try:
                        time.sleep(3)
                        driver.find_elements_by_xpath('//div[@class="user-display__author"]/h1[@class="user-display__name"]/a')[noneInc].click()
                        time.sleep(3)

                        for badge in driver.find_elements_by_class_name('badges__badge'):
                            if badge.get_attribute("alt") in self.badges:
                                followBtn = driver.find_element_by_xpath('//div[@class="user-card__buttons"]/button')
                                if(followBtn.get_attribute("title")[ 0 : 6 ] == "Follow"):

                                    self.userNames.append(driver.find_element_by_xpath('//div[@class="user-card__user-details"]/h1[@class="name"]').text)

                                    followBtn.click()

                                    time.sleep(3)

                                    # self.createExcelForFollowing()

                                    time.sleep(1)



                    except WebDriverException:
                        print("Is not clickable")


                    
                    
                    noneInc = noneInc + 1

        time.sleep(2)

        self.createExcelForFollowing()

        # wait to close browser
        time.sleep(3)


    def createExcelForFollowing(self):
        now = datetime.now()
        current_time = now.strftime("%d-%b")
        if(os.path.exists(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")):
            if (os.access(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx",os.R_OK)):
                wrkbk = openpyxl.load_workbook(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")
                sh = wrkbk.active
                newEmptyRow = sh.max_row + 1
                wrkbk.save(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")
                wrkbk.close()
            else:
                print("File permision denied: Follow module")
        else:
            book = openpyxl.Workbook()
            book.save(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")
            book.close()
            newEmptyRow = 1

        book = openpyxl.load_workbook(filename=os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")
        sheet = book.active
        # ws = book.worksheets[0]

        userNamesForExcel = list(set(self.userNames))
        userNamesForExcel = list(filter(None, userNamesForExcel))

        # for item in self.userNames :
        newEmptyRow = newEmptyRow - 1
        for item in userNamesForExcel :
            newEmptyRow = newEmptyRow + 1
            sheet.cell(row=newEmptyRow, column=1).value = item

        book.save(os.path.abspath("archive") + "\\" + "following"+current_time+".xlsx")
        book.close()
        

    def __init__(self, status, hashtag, badge, parlerUsername, parlerPassowrd):
        self.progstatus = status
        self.hashtag = hashtag
        self.badges = badge
        self.parlerUsername = parlerUsername
        self.parlerPassword = parlerPassowrd
        self.userNames = []
        
        # self.parlerSelenium()
        t1 = threading.Thread(target=self.parlerSelenium);
        t1.start()

        t1.join()
        print("Following module completed")

